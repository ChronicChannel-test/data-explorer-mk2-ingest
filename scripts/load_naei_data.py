#!/usr/bin/env python3
"""Local extractor/loader for NAEI PV PivotTableViewer workbook data.

Commands:
- extract-pv-xlsx: read visible pivot sheets and write normalized CSVs
- load-pv: load normalized CSVs into Supabase/Postgres
- run-pv-ingest: run extract + load in one step
"""

from __future__ import annotations

import argparse
import csv
import math
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Protocol, Sequence, Set, Tuple

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover - exercised in environments without dependency
    load_dotenv = None

try:
    import psycopg
    from psycopg import Connection
except ImportError:  # pragma: no cover - exercised in environments without dependency
    psycopg = None
    Connection = Any  # type: ignore[misc,assignment]


PREFIX_PATTERN = re.compile(r"^naei(\d{4})(ds|pv)$", re.IGNORECASE)
DEFAULT_OUTPUT_DIR = Path.home() / "naei-ingest-output"

VISIBLE_PV_HEADER_ROW = 14
VISIBLE_PV_FIXED_HEADERS = ("NFRCode", "SourceName", "ActivityName", "Emission Unit")


@dataclass(frozen=True)
class VisiblePVSheetSpec:
    subtype: str
    pollutant_header: str


TARGET_PV_SHEETS: Dict[str, VisiblePVSheetSpec] = {
    "AirPollutants":     VisiblePVSheetSpec(subtype="AQ",  pollutant_header="Pollutant"),
    "HeavyMetals":       VisiblePVSheetSpec(subtype="HM",  pollutant_header="Pollutant"),
    "ParticulateMatter": VisiblePVSheetSpec(subtype="PM",  pollutant_header="Particle Size"),
    "POPs&PAHs":         VisiblePVSheetSpec(subtype="POP", pollutant_header="Pollutant"),
}

NORMALIZED_COLUMNS: List[str] = [
    "pollutant",
    "reporting_year",
    "emission_unit",
    "source_name",
    "activity_name",
    "emission_value",
    "nfr_code",
]


class SubparserAction(Protocol):
    def add_parser(self, name: str, *, help: Optional[str] = None, **kwargs: Any) -> argparse.ArgumentParser:
        ...


@dataclass(frozen=True)
class DatasetMeta:
    dataset_prefix: str
    file_name: str
    source_url: Optional[str]
    extracted_at: datetime


@dataclass(frozen=True)
class NormalizedPVRow:
    extracted_at: str
    source_sheet: str
    dataset_prefix: str
    pollutant: str
    reporting_year: int
    emission_unit: Optional[str]
    source_name: str
    activity_name: str
    emission_value: float
    nfr_code: str

    def series_lookup_key(self) -> Tuple[str, str, str, str]:
        return (
            self.pollutant.lower(),
            self.nfr_code.lower(),
            self.source_name.lower(),
            self.activity_name.lower(),
        )


@dataclass
class ExtractSheetSummary:
    sheet_name: str
    csv_path: Path
    rows_written: int = 0
    rows_skipped: int = 0
    skipped_reasons: Dict[str, int] = field(default_factory=lambda: defaultdict(int))


@dataclass
class ExtractRunSummary:
    workbook_path: Path
    dataset_prefix: str
    output_dir: Path
    run_timestamp: datetime
    sheets: List[ExtractSheetSummary]


@dataclass
class LoadFileSummary:
    csv_path: Path
    dataset_prefix: str
    dataset_file_id: int
    rows_read: int
    rows_loaded: int
    rows_skipped: int
    skipped_reasons: Dict[str, int]


@dataclass
class LoadRunSummary:
    csv_paths: List[Path]
    files: List[LoadFileSummary]
    unit_conflicts: List[str]
    lookup_delta: Dict[str, int]


def execute_values(cur: Any, query: str, rows: Sequence[Sequence[object]], page_size: int = 1000) -> None:
    """Minimal execute_values helper to keep dependencies small."""

    if not rows:
        return
    if "%s" not in query:
        raise ValueError("Query must include a VALUES %s placeholder")

    num_columns = len(rows[0])
    value_template = "(" + ",".join(["%s"] * num_columns) + ")"
    for start in range(0, len(rows), page_size):
        chunk = rows[start : start + page_size]
        placeholders = ", ".join([value_template] * len(chunk))
        flat_params: List[object] = []
        for row in chunk:
            if len(row) != num_columns:
                raise ValueError("All rows must have same column count")
            flat_params.extend(row)
        cur.execute(query.replace("VALUES %s", f"VALUES {placeholders}"), flat_params)


def normalize_dataset_prefix(value: str) -> str:
    value = value.strip()
    match = PREFIX_PATTERN.match(value)
    if not match:
        return value
    year, suffix = match.groups()
    return f"NAEI{year}{suffix.lower()}"


def infer_dataset_prefix_from_csv_name(file_name: str) -> Optional[str]:
    stem = Path(file_name).stem
    if "_" not in stem:
        return None
    prefix = stem.split("_", 1)[0].strip()
    if not prefix:
        return None
    try:
        return require_pv_dataset_prefix(prefix)
    except ValueError:
        return None


def infer_source_sheet_from_csv_name(file_name: str) -> Optional[str]:
    stem = Path(file_name).stem
    parts = [part.strip() for part in stem.split("_") if part.strip()]
    if len(parts) >= 3:
        return parts[2]
    return None


def require_pv_dataset_prefix(value: str) -> str:
    normalized = normalize_dataset_prefix(value)
    if not PREFIX_PATTERN.match(normalized) or not normalized.lower().endswith("pv"):
        raise ValueError(
            f"Invalid dataset prefix '{value}'. Expected NAEI####pv format."
        )
    return normalized


def load_env_dsn(passed_dsn: Optional[str] = None) -> str:
    if passed_dsn:
        return passed_dsn
    if load_dotenv is not None:
        load_dotenv()
    dsn = os.environ.get("SUPABASE_DB_URL") or os.environ.get("DATABASE_URL")
    if not dsn:
        raise RuntimeError("Missing DSN. Pass --dsn or set SUPABASE_DB_URL / DATABASE_URL")
    return dsn


def require_psycopg() -> None:
    if psycopg is None:
        raise RuntimeError("psycopg is not installed. Install dependencies from requirements.txt")


def clean_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_year(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, int):
        year = value
    elif isinstance(value, float):
        if not math.isfinite(value):
            return None
        year = int(value)
    else:
        text = str(value).strip()
        if not text:
            return None
        if text.endswith(".0"):
            text = text[:-2]
        if not text.isdigit():
            return None
        year = int(text)
    if year < 1900 or year > 2100:
        return None
    return year


def parse_year_arg(value: str) -> int:
    year = parse_year(value)
    if year is None:
        raise argparse.ArgumentTypeError("Year must be an integer between 1900 and 2100")
    return year


def parse_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        numeric = float(value)
        if not math.isfinite(numeric):
            return None
        return numeric
    text = str(value).strip()
    if not text or text == "-":
        return None
    text = text.replace(",", "")
    try:
        numeric = float(text)
    except ValueError:
        return None
    if not math.isfinite(numeric):
        return None
    return numeric


def parse_excel_timestamp(value: Any) -> Optional[str]:
    """Convert Excel timestamp values to ISO8601 strings when possible."""

    if value is None:
        return None

    if isinstance(value, datetime):
        dt = value if value.tzinfo is not None else value.replace(tzinfo=timezone.utc)
        return dt.isoformat()

    if isinstance(value, (int, float)):
        if not math.isfinite(float(value)):
            return None
        dt = (datetime(1899, 12, 30) + timedelta(days=float(value))).replace(tzinfo=timezone.utc)
        return dt.isoformat()

    text = str(value).strip()
    if not text:
        return None

    try:
        numeric = float(text)
    except ValueError:
        numeric = None

    if numeric is not None and math.isfinite(numeric):
        dt = (datetime(1899, 12, 30) + timedelta(days=float(numeric))).replace(tzinfo=timezone.utc)
        return dt.isoformat()

    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(text).isoformat()
    except ValueError:
        return text


def validate_visible_sheet_headers(
    sheet_name: str,
    sheet_spec: VisiblePVSheetSpec,
    headers: Sequence[Any],
) -> Tuple[Dict[str, int], List[Tuple[int, int]]]:
    normalized = [str(value).strip() if value is not None else "" for value in headers]
    header_index = {name: idx for idx, name in enumerate(normalized) if name}

    required = (sheet_spec.pollutant_header,) + VISIBLE_PV_FIXED_HEADERS
    missing = [name for name in required if name not in header_index]
    if missing:
        raise ValueError(
            f"Unexpected visible-sheet headers in {sheet_name}. Missing required columns: {', '.join(missing)}"
        )

    year_columns: List[Tuple[int, int]] = []
    for idx, value in enumerate(headers):
        year = parse_year(value)
        if year is not None:
            year_columns.append((idx, year))

    if not year_columns:
        raise ValueError(f"{sheet_name} has no parseable year columns in header row {VISIBLE_PV_HEADER_ROW}")

    return header_index, year_columns


def extract_visible_sheet_timestamp(worksheet: Any, fallback_extracted_at: str) -> str:
    for values in worksheet.iter_rows(min_row=1, max_row=VISIBLE_PV_HEADER_ROW - 1, values_only=True):
        label = clean_text(values[1] if len(values) > 1 else None)
        if label and label.lower() == "time-stamp":
            raw = clean_text(values[2] if len(values) > 2 else None)
            if raw is None or raw == "(All)":
                return fallback_extracted_at
            return parse_excel_timestamp(raw) or fallback_extracted_at
    return fallback_extracted_at


def is_grand_total_row(*values: Optional[str]) -> bool:
    for value in values:
        if value is not None and value.strip().lower() == "grand total":
            return True
    return False


def normalized_row_to_csv(row: NormalizedPVRow) -> Dict[str, str]:
    return {
        "pollutant": row.pollutant,
        "reporting_year": str(row.reporting_year),
        "emission_unit": row.emission_unit or "",
        "source_name": row.source_name,
        "activity_name": row.activity_name,
        "emission_value": format(row.emission_value, ".15g"),
        "nfr_code": row.nfr_code,
    }


def parse_normalized_csv_row(
    raw_row: Mapping[str, str],
    dataset_prefix_override: Optional[str],
    dataset_prefix_fallback: Optional[str],
    source_sheet_fallback: Optional[str],
    extracted_at_fallback: Optional[str],
    force_reporting_year_override: Optional[int] = None,
) -> Tuple[Optional[NormalizedPVRow], Optional[str]]:
    for col in NORMALIZED_COLUMNS:
        if col not in raw_row:
            return None, f"missing_column_{col}"

    dataset_prefix = (
        require_pv_dataset_prefix(dataset_prefix_override)
        if dataset_prefix_override
        else clean_text(raw_row.get("dataset_prefix")) or dataset_prefix_fallback
    )
    if not dataset_prefix:
        return None, "missing_dataset_prefix"
    if not PREFIX_PATTERN.match(dataset_prefix) or not dataset_prefix.lower().endswith("pv"):
        return None, "invalid_dataset_prefix"

    pollutant = clean_text(raw_row.get("pollutant"))
    source_name = clean_text(raw_row.get("source_name"))
    activity_name = clean_text(raw_row.get("activity_name"))
    nfr_code = clean_text(raw_row.get("nfr_code"))
    reporting_year = (
        force_reporting_year_override
        if force_reporting_year_override is not None
        else parse_year(raw_row.get("reporting_year"))
    )
    emission_value = parse_float(raw_row.get("emission_value"))
    source_sheet = clean_text(raw_row.get("source_sheet")) or source_sheet_fallback or "unknown"
    emission_unit = clean_text(raw_row.get("emission_unit"))

    if not pollutant or reporting_year is None:
        return None, "missing_pollutant_or_year"
    if source_name is None:
        return None, "missing_source"
    if activity_name is None:
        return None, "missing_activity"
    if nfr_code is None:
        return None, "missing_nfr_code"
    if emission_value is None:
        return None, "missing_or_invalid_emission"

    extracted_at = clean_text(raw_row.get("extracted_at")) or extracted_at_fallback or ""
    row = NormalizedPVRow(
        extracted_at=extracted_at,
        source_sheet=source_sheet,
        dataset_prefix=dataset_prefix,
        pollutant=pollutant,
        reporting_year=reporting_year,
        emission_unit=emission_unit,
        source_name=source_name,
        activity_name=activity_name,
        emission_value=emission_value,
        nfr_code=nfr_code,
    )
    return row, None


def gather_csv_paths(path: Path) -> List[Path]:
    if path.is_file():
        return [path]
    if path.is_dir():
        return sorted(
            item for item in path.rglob("*.csv")
            if item.is_file() and re.match(r"naei\d{4}pv_[A-Z]+\.csv$", item.name, re.IGNORECASE)
        )
    raise FileNotFoundError(path)


def dedupe_csv_paths_by_name(csv_paths: Sequence[Path]) -> Tuple[List[Path], Dict[str, List[Path]]]:
    grouped: Dict[str, List[Path]] = defaultdict(list)
    for csv_path in csv_paths:
        grouped[csv_path.name].append(csv_path)

    duplicates = {name: paths for name, paths in grouped.items() if len(paths) > 1}
    selected: List[Path] = []
    for name, paths in grouped.items():
        best = max(paths, key=lambda p: (p.stat().st_mtime, str(p)))
        selected.append(best)
    return sorted(selected), duplicates


def extract_pv_xlsx(
    xlsx_path: Path,
    output_dir: Path,
    dataset_prefix: str,
    force_reporting_year: Optional[int] = None,
) -> ExtractRunSummary:
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)
    if force_reporting_year is not None:
        raise ValueError(
            "--force-reporting-year is not supported for visible-sheet extraction. "
            "Use load-pv --force-reporting-year on normalized CSVs instead."
        )

    normalized_prefix = require_pv_dataset_prefix(dataset_prefix)

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency-driven
        raise RuntimeError("openpyxl is not installed. Install dependencies from requirements.txt") from exc

    output_dir.mkdir(parents=True, exist_ok=True)
    run_timestamp = datetime.now(timezone.utc).replace(microsecond=0)
    fallback_timestamp = run_timestamp.isoformat()

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    missing_sheets = sorted(set(TARGET_PV_SHEETS) - set(workbook.sheetnames))
    if missing_sheets:
        raise ValueError(f"Workbook is missing required sheets: {', '.join(missing_sheets)}")

    summaries: List[ExtractSheetSummary] = []

    try:
        for sheet_name, sheet_spec in TARGET_PV_SHEETS.items():
            worksheet = workbook[sheet_name]
            header_row = next(
                worksheet.iter_rows(
                    min_row=VISIBLE_PV_HEADER_ROW,
                    max_row=VISIBLE_PV_HEADER_ROW,
                    values_only=True,
                ),
                None,
            )
            if header_row is None:
                raise ValueError(f"{sheet_name} has no header row at row {VISIBLE_PV_HEADER_ROW}")

            header_index, year_columns = validate_visible_sheet_headers(sheet_name, sheet_spec, header_row)
            extracted_at = extract_visible_sheet_timestamp(worksheet, fallback_timestamp)

            csv_name = f"{normalized_prefix}_{sheet_spec.subtype}.csv"
            csv_path = output_dir / csv_name
            sheet_summary = ExtractSheetSummary(sheet_name=sheet_name, csv_path=csv_path)

            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(handle, fieldnames=NORMALIZED_COLUMNS)
                writer.writeheader()

                for values in worksheet.iter_rows(min_row=VISIBLE_PV_HEADER_ROW + 1, values_only=True):
                    pollutant = clean_text(
                        values[header_index[sheet_spec.pollutant_header]]
                        if header_index[sheet_spec.pollutant_header] < len(values)
                        else None
                    )
                    nfr_code = clean_text(values[header_index["NFRCode"]] if header_index["NFRCode"] < len(values) else None)
                    source_name = clean_text(
                        values[header_index["SourceName"]]
                        if header_index["SourceName"] < len(values)
                        else None
                    )
                    activity_name = clean_text(
                        values[header_index["ActivityName"]]
                        if header_index["ActivityName"] < len(values)
                        else None
                    )
                    emission_unit = clean_text(
                        values[header_index["Emission Unit"]]
                        if header_index["Emission Unit"] < len(values)
                        else None
                    )

                    if is_grand_total_row(pollutant, nfr_code, source_name, activity_name):
                        sheet_summary.rows_skipped += 1
                        sheet_summary.skipped_reasons["summary_row"] += 1
                        continue

                    if not pollutant or nfr_code is None or source_name is None or activity_name is None:
                        sheet_summary.rows_skipped += 1
                        sheet_summary.skipped_reasons["missing_dimension_values"] += 1
                        continue

                    row_wrote_any = False
                    for year_index, year in year_columns:
                        reporting_year = year
                        raw_emission = values[year_index] if year_index < len(values) else None
                        emission_value = parse_float(raw_emission)
                        if emission_value is None:
                            if clean_text(raw_emission) is not None:
                                sheet_summary.skipped_reasons["invalid_emission_cell"] += 1
                            continue

                        normalized_row = NormalizedPVRow(
                            extracted_at=extracted_at,
                            source_sheet=sheet_name,
                            dataset_prefix=normalized_prefix,
                            pollutant=pollutant,
                            reporting_year=reporting_year,
                            emission_unit=emission_unit,
                            source_name=source_name,
                            activity_name=activity_name,
                            emission_value=emission_value,
                            nfr_code=nfr_code,
                        )
                        writer.writerow(normalized_row_to_csv(normalized_row))
                        sheet_summary.rows_written += 1
                        row_wrote_any = True

                    if not row_wrote_any:
                        sheet_summary.rows_skipped += 1
                        sheet_summary.skipped_reasons["missing_or_invalid_emission"] += 1
                        continue

            summaries.append(sheet_summary)
    finally:
        workbook.close()

    return ExtractRunSummary(
        workbook_path=xlsx_path,
        dataset_prefix=normalized_prefix,
        output_dir=output_dir,
        run_timestamp=run_timestamp,
        sheets=summaries,
    )


class DimensionCache:
    """Dimension lookup cache with created/reused counters."""

    def __init__(self) -> None:
        self.pollutant: Dict[str, int] = {}
        self.pollutant_unit: Dict[int, Optional[str]] = {}
        self.pollutant_label: Dict[int, str] = {}
        self.nfr: Dict[str, int] = {}
        self.source: Dict[str, int] = {}
        self.activity: Dict[str, int] = {}
        self.unit: Dict[str, int] = {}
        self.unit_conflicts: Set[str] = set()
        self.stats: Dict[str, int] = defaultdict(int)

    @staticmethod
    def normalize_value(value: Optional[str]) -> Optional[str]:
        if value is None:
            return None
        value = value.strip()
        return value or None

    def _track(self, name: str) -> None:
        self.stats[name] += 1

    def _pollutant_id_from_alias(self, cur: Any, alias_key: str) -> Optional[Tuple[int, Optional[str], Optional[str]]]:
        cur.execute(
            """
            SELECT p.id, p.emission_unit, p.pollutant
            FROM naei_global_t_pollutant_alias a
            JOIN naei_global_t_pollutant p ON p.id = a.pollutant_id
            WHERE a.alias_key = %s
            """,
            (alias_key,),
        )
        row = cur.fetchone()
        if row is None:
            return None
        return int(row[0]), self.normalize_value(row[1]), row[2]

    def _upsert_pollutant_alias(self, cur: Any, alias_name: str, pollutant_id: int) -> None:
        cur.execute(
            """
            INSERT INTO naei_global_t_pollutant_alias (alias_name, pollutant_id)
            VALUES (%s, %s)
            ON CONFLICT (alias_key)
            DO UPDATE SET pollutant_id = EXCLUDED.pollutant_id,
                          alias_name = EXCLUDED.alias_name
            """,
            (alias_name, pollutant_id),
        )

    def _apply_pollutant_unit_rule(
        self,
        cur: Any,
        pollutant_id: int,
        pollutant_label: str,
        incoming_unit: Optional[str],
    ) -> None:
        incoming = self.normalize_value(incoming_unit)
        existing = self.pollutant_unit.get(pollutant_id)
        if incoming is None:
            return

        if existing is None:
            cur.execute(
                """
                UPDATE naei_global_t_pollutant
                SET emission_unit = %s
                WHERE id = %s AND emission_unit IS NULL
                """,
                (incoming, pollutant_id),
            )
            if cur.rowcount:
                self.pollutant_unit[pollutant_id] = incoming
                return
            cur.execute("SELECT emission_unit FROM naei_global_t_pollutant WHERE id = %s", (pollutant_id,))
            row = cur.fetchone()
            self.pollutant_unit[pollutant_id] = self.normalize_value(row[0]) if row else incoming
            return

        if existing.lower() != incoming.lower():
            warning = (
                f"{pollutant_label}: keeping canonical unit '{existing}' "
                f"and ignoring conflicting unit '{incoming}'"
            )
            self.unit_conflicts.add(warning)

    def pollutant_id(self, cur: Any, name: str, emission_unit: Optional[str]) -> int:
        normalized = self.normalize_value(name)
        if not normalized:
            raise ValueError("Pollutant cannot be blank")
        key = normalized.lower()
        if key in self.pollutant:
            pollutant_id = self.pollutant[key]
            self._track("pollutant_cache_hit")
            self._apply_pollutant_unit_rule(cur, pollutant_id, normalized, emission_unit)
            return pollutant_id

        alias_row = self._pollutant_id_from_alias(cur, key)
        if alias_row is not None:
            pollutant_id, canonical_unit, canonical_label = alias_row
            self.pollutant[key] = pollutant_id
            self.pollutant_unit[pollutant_id] = canonical_unit
            self.pollutant_label[pollutant_id] = canonical_label or normalized
            self._track("pollutant_reused")
            self._upsert_pollutant_alias(cur, normalized, pollutant_id)
            self._apply_pollutant_unit_rule(cur, pollutant_id, canonical_label or normalized, emission_unit)
            return pollutant_id

        cur.execute(
            """
            SELECT id, emission_unit, pollutant
            FROM naei_global_t_pollutant
            WHERE lower(pollutant) = lower(%s)
            LIMIT 1
            """,
            (normalized,),
        )
        row = cur.fetchone()
        if row is None:
            cur.execute(
                """
                INSERT INTO naei_global_t_pollutant (pollutant, emission_unit)
                VALUES (%s, %s)
                RETURNING id, emission_unit, pollutant
                """,
                (normalized, self.normalize_value(emission_unit)),
            )
            row = cur.fetchone()
            self._track("pollutant_created")
        else:
            self._track("pollutant_reused")

        pollutant_id = int(row[0])
        canonical_unit = self.normalize_value(row[1])
        canonical_label = row[2] or normalized
        self.pollutant[key] = pollutant_id
        self.pollutant_unit[pollutant_id] = canonical_unit
        self.pollutant_label[pollutant_id] = canonical_label
        self._upsert_pollutant_alias(cur, normalized, pollutant_id)
        self._apply_pollutant_unit_rule(cur, pollutant_id, canonical_label, emission_unit)
        return pollutant_id

    def nfr_id(self, cur: Any, code: str) -> int:
        normalized = self.normalize_value(code)
        if not normalized:
            raise ValueError("NFR code cannot be blank")
        key = normalized.lower()
        if key in self.nfr:
            self._track("nfr_cache_hit")
            return self.nfr[key]

        cur.execute(
            """
            SELECT id
            FROM naei_global_t_nfrcode
            WHERE lower(nfr_code) = lower(%s)
            LIMIT 1
            """,
            (normalized,),
        )
        row = cur.fetchone()
        if row is None:
            cur.execute(
                "INSERT INTO naei_global_t_nfrcode (nfr_code, description) VALUES (%s, %s) RETURNING id",
                (normalized, normalized),
            )
            row = cur.fetchone()
            self._track("nfr_created")
        else:
            self._track("nfr_reused")

        nfr_id = int(row[0])
        self.nfr[key] = nfr_id
        return nfr_id

    def source_id(self, cur: Any, name: str) -> int:
        normalized = self.normalize_value(name)
        if not normalized:
            raise ValueError("Source cannot be blank")
        key = normalized.lower()
        if key in self.source:
            self._track("source_cache_hit")
            return self.source[key]

        cur.execute(
            """
            SELECT id
            FROM naei_global_t_sourcename
            WHERE lower(source_name) = lower(%s)
            LIMIT 1
            """,
            (normalized,),
        )
        row = cur.fetchone()
        if row is None:
            cur.execute(
                "INSERT INTO naei_global_t_sourcename (source_name) VALUES (%s) RETURNING id",
                (normalized,),
            )
            row = cur.fetchone()
            self._track("source_created")
        else:
            self._track("source_reused")

        source_id = int(row[0])
        self.source[key] = source_id
        return source_id

    def activity_id(self, cur: Any, name: str) -> int:
        normalized = self.normalize_value(name)
        if not normalized:
            raise ValueError("Activity cannot be blank")
        key = normalized.lower()
        if key in self.activity:
            self._track("activity_cache_hit")
            return self.activity[key]

        cur.execute(
            """
            SELECT id
            FROM naei_global_t_activityname
            WHERE lower(activity_name) = lower(%s)
            LIMIT 1
            """,
            (normalized,),
        )
        row = cur.fetchone()
        if row is None:
            cur.execute(
                "INSERT INTO naei_global_t_activityname (activity_name) VALUES (%s) RETURNING id",
                (normalized,),
            )
            row = cur.fetchone()
            self._track("activity_created")
        else:
            self._track("activity_reused")

        activity_id = int(row[0])
        self.activity[key] = activity_id
        return activity_id

    def unit_id(self, cur: Any, name: Optional[str]) -> Optional[int]:
        normalized = self.normalize_value(name)
        if not normalized:
            return None
        key = normalized.lower()
        if key in self.unit:
            self._track("unit_cache_hit")
            return self.unit[key]

        cur.execute(
            """
            SELECT unit_id
            FROM unit
            WHERE lower(unit_name) = lower(%s)
            LIMIT 1
            """,
            (normalized,),
        )
        row = cur.fetchone()
        if row is None:
            cur.execute("INSERT INTO unit (unit_name) VALUES (%s) RETURNING unit_id", (normalized,))
            row = cur.fetchone()
            self._track("unit_created")
        else:
            self._track("unit_reused")

        unit_id = int(row[0])
        self.unit[key] = unit_id
        return unit_id


class NAEIPVLoader:
    def __init__(self, conn: Connection) -> None:
        self.conn = conn
        self.dim_cache = DimensionCache()

    def ensure_dataset_file(self, cur: Any, meta: DatasetMeta) -> int:
        cur.execute(
            """
            INSERT INTO dataset_file (dataset_prefix, file_name, extracted_at, source_url)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (dataset_prefix, file_name)
            DO UPDATE SET extracted_at = EXCLUDED.extracted_at,
                          source_url = COALESCE(EXCLUDED.source_url, dataset_file.source_url)
            RETURNING dataset_file_id
            """,
            (meta.dataset_prefix, meta.file_name, meta.extracted_at, meta.source_url),
        )
        return int(cur.fetchone()[0])

    def upsert_pv_series(
        self,
        cur: Any,
        *,
        dataset_file_id: int,
        pollutant_id: int,
        nfr_group_id: int,
        source_id: int,
        activity_id: int,
    ) -> int:
        cur.execute(
            """
            SELECT min(pv_series_id)
            FROM naei2024pv_series
            WHERE dataset_file_id = %s
              AND pollutant_id = %s
              AND nfr_group_id = %s
              AND source_id = %s
              AND activity_id = %s
            """,
            (dataset_file_id, pollutant_id, nfr_group_id, source_id, activity_id),
        )
        existing = cur.fetchone()[0]
        if existing is not None:
            return int(existing)

        cur.execute(
            """
            INSERT INTO naei2024pv_series (
                dataset_file_id,
                pollutant_id,
                nfr_group_id,
                source_id,
                activity_id
            )
            VALUES (%s, %s, %s, %s, %s)
            RETURNING pv_series_id
            """,
            (dataset_file_id, pollutant_id, nfr_group_id, source_id, activity_id),
        )
        return int(cur.fetchone()[0])

    def _flush_values(self, cur: Any, rows: List[Tuple[int, int, str, float]]) -> None:
        aggregated: Dict[Tuple[int, int, str], float] = defaultdict(float)
        for pv_series_id, reporting_year, metric_label, metric_value in rows:
            aggregated[(pv_series_id, reporting_year, metric_label)] += metric_value

        payload: List[Tuple[int, int, str, float]] = [
            (pv_series_id, reporting_year, metric_label, metric_value)
            for (pv_series_id, reporting_year, metric_label), metric_value in aggregated.items()
        ]
        execute_values(
            cur,
            """
            INSERT INTO naei2024pv_values (pv_series_id, reporting_year, metric_label, metric_value)
            VALUES %s
            ON CONFLICT (pv_series_id, reporting_year, metric_label)
            DO UPDATE SET metric_value = EXCLUDED.metric_value
            """,
            payload,
            page_size=5000,
        )

    def load_pv_csv(
        self,
        csv_path: Path,
        *,
        dataset_prefix_override: Optional[str] = None,
        source_url: Optional[str] = None,
        force_reporting_year_override: Optional[int] = None,
    ) -> LoadFileSummary:
        extracted_at = datetime.fromtimestamp(csv_path.stat().st_mtime, tz=timezone.utc)
        rows_read = 0
        rows_loaded = 0
        rows_skipped = 0
        skipped_reasons: Dict[str, int] = defaultdict(int)

        with self.conn.cursor() as cur, csv_path.open("r", newline="", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            if reader.fieldnames is None:
                raise ValueError(f"Missing header row in {csv_path}")
            missing_cols = [col for col in NORMALIZED_COLUMNS if col not in reader.fieldnames]
            if missing_cols:
                raise ValueError(f"{csv_path.name} is missing required columns: {', '.join(missing_cols)}")

            dataset_file_id: Optional[int] = None
            inferred_prefix = infer_dataset_prefix_from_csv_name(csv_path.name)
            active_prefix: Optional[str] = (
                require_pv_dataset_prefix(dataset_prefix_override)
                if dataset_prefix_override
                else inferred_prefix
            )
            source_sheet_fallback = infer_source_sheet_from_csv_name(csv_path.name) or csv_path.stem
            extracted_at_fallback = extracted_at.isoformat()
            batch_rows: List[Tuple[int, int, str, float]] = []

            for raw_row in reader:
                rows_read += 1
                normalized_row, reason = parse_normalized_csv_row(
                    raw_row,
                    dataset_prefix_override,
                    dataset_prefix_fallback=active_prefix,
                    source_sheet_fallback=source_sheet_fallback,
                    extracted_at_fallback=extracted_at_fallback,
                    force_reporting_year_override=force_reporting_year_override,
                )
                if normalized_row is None:
                    rows_skipped += 1
                    if reason:
                        skipped_reasons[reason] += 1
                    continue

                if active_prefix is None:
                    active_prefix = require_pv_dataset_prefix(normalized_row.dataset_prefix)

                if dataset_file_id is None:
                    meta = DatasetMeta(
                        dataset_prefix=active_prefix,
                        file_name=csv_path.name,
                        source_url=source_url,
                        extracted_at=extracted_at,
                    )
                    dataset_file_id = self.ensure_dataset_file(cur, meta)

                pollutant_id = self.dim_cache.pollutant_id(cur, normalized_row.pollutant, normalized_row.emission_unit)
                nfr_id = self.dim_cache.nfr_id(cur, normalized_row.nfr_code)
                source_id = self.dim_cache.source_id(cur, normalized_row.source_name)
                activity_id = self.dim_cache.activity_id(cur, normalized_row.activity_name)
                self.dim_cache.unit_id(cur, normalized_row.emission_unit)

                pv_series_id = self.upsert_pv_series(
                    cur,
                    dataset_file_id=dataset_file_id,
                    pollutant_id=pollutant_id,
                    nfr_group_id=nfr_id,
                    source_id=source_id,
                    activity_id=activity_id,
                )

                batch_rows.append((pv_series_id, normalized_row.reporting_year, "value", normalized_row.emission_value))
                rows_loaded += 1

                if len(batch_rows) >= 5000:
                    self._flush_values(cur, batch_rows)
                    batch_rows.clear()

            if batch_rows:
                self._flush_values(cur, batch_rows)

            if dataset_file_id is None:
                raise ValueError(f"{csv_path.name} produced no valid rows to load")

            self.conn.commit()

        final_prefix = active_prefix
        if final_prefix is None:
            raise ValueError(
                f"Could not determine dataset_prefix for {csv_path.name}. "
                "Pass --dataset-prefix or use file names prefixed like NAEI2024pv_..."
            )

        return LoadFileSummary(
            csv_path=csv_path,
            dataset_prefix=final_prefix,
            dataset_file_id=dataset_file_id,
            rows_read=rows_read,
            rows_loaded=rows_loaded,
            rows_skipped=rows_skipped,
            skipped_reasons=dict(sorted(skipped_reasons.items())),
        )


def run_load_pv(
    *,
    path: Path,
    dsn: str,
    source_url: Optional[str],
    dataset_prefix_override: Optional[str],
    force_reporting_year_override: Optional[int] = None,
) -> LoadRunSummary:
    csv_paths = gather_csv_paths(path)
    if not csv_paths:
        raise RuntimeError(f"No CSV files found under {path}")
    csv_paths, duplicates = dedupe_csv_paths_by_name(csv_paths)
    if duplicates:
        print(
            "Warning: duplicate CSV filenames detected; using newest file per name and ignoring older duplicates.",
            file=sys.stderr,
        )
        for name in sorted(duplicates):
            ignored = sorted(str(path) for path in duplicates[name][:-1])
            for path_str in ignored:
                print(f"  ignored duplicate: {path_str}", file=sys.stderr)
    return run_load_pv_paths(
        csv_paths=csv_paths,
        dsn=dsn,
        source_url=source_url,
        dataset_prefix_override=dataset_prefix_override,
        force_reporting_year_override=force_reporting_year_override,
    )


def run_load_pv_paths(
    *,
    csv_paths: Sequence[Path],
    dsn: str,
    source_url: Optional[str],
    dataset_prefix_override: Optional[str],
    force_reporting_year_override: Optional[int] = None,
) -> LoadRunSummary:
    require_psycopg()
    deduped_paths, duplicates = dedupe_csv_paths_by_name(list(csv_paths))
    if duplicates:
        print(
            "Warning: duplicate CSV filenames detected; using newest file per name and ignoring older duplicates.",
            file=sys.stderr,
        )
        for name in sorted(duplicates):
            ignored = sorted(str(path) for path in duplicates[name][:-1])
            for path_str in ignored:
                print(f"  ignored duplicate: {path_str}", file=sys.stderr)
    with psycopg.connect(dsn) as conn:
        loader = NAEIPVLoader(conn)
        files: List[LoadFileSummary] = []
        before_stats = dict(loader.dim_cache.stats)
        before_conflicts = set(loader.dim_cache.unit_conflicts)

        for csv_path in deduped_paths:
            files.append(
                loader.load_pv_csv(
                    csv_path,
                    dataset_prefix_override=dataset_prefix_override,
                    source_url=source_url,
                    force_reporting_year_override=force_reporting_year_override,
                )
            )

        unit_conflicts = sorted(loader.dim_cache.unit_conflicts - before_conflicts)
        lookup_delta: Dict[str, int] = {}
        for key, after_value in loader.dim_cache.stats.items():
            lookup_delta[key] = after_value - before_stats.get(key, 0)

    return LoadRunSummary(
        csv_paths=deduped_paths,
        files=files,
        unit_conflicts=unit_conflicts,
        lookup_delta=dict(sorted(lookup_delta.items())),
    )


def print_extract_summary(summary: ExtractRunSummary) -> None:
    print(f"Workbook: {summary.workbook_path}")
    print(f"Dataset prefix: {summary.dataset_prefix}")
    print(f"Output directory: {summary.output_dir}")
    print(f"Extraction run timestamp: {summary.run_timestamp.isoformat()}")
    print("Extracted CSVs:")
    for sheet in summary.sheets:
        print(
            f"- {sheet.sheet_name}: {sheet.csv_path.name} "
            f"(written={sheet.rows_written}, skipped={sheet.rows_skipped})"
        )
        for reason, count in sorted(sheet.skipped_reasons.items()):
            print(f"  skip[{reason}]={count}")


def print_load_summary(summary: LoadRunSummary) -> None:
    total_read = sum(item.rows_read for item in summary.files)
    total_loaded = sum(item.rows_loaded for item in summary.files)
    total_skipped = sum(item.rows_skipped for item in summary.files)

    print("Load summary:")
    for file_summary in summary.files:
        print(
            f"- {file_summary.csv_path.name}: "
            f"dataset_file_id={file_summary.dataset_file_id}, "
            f"rows_read={file_summary.rows_read}, "
            f"rows_loaded={file_summary.rows_loaded}, "
            f"rows_skipped={file_summary.rows_skipped}"
        )
        for reason, count in file_summary.skipped_reasons.items():
            print(f"  skip[{reason}]={count}")

    print(
        f"Totals: files={len(summary.files)}, rows_read={total_read}, "
        f"rows_loaded={total_loaded}, rows_skipped={total_skipped}"
    )

    if summary.lookup_delta:
        print("Lookup activity:")
        for key, value in summary.lookup_delta.items():
            if value:
                print(f"- {key}: {value}")

    if summary.unit_conflicts:
        print("Pollutant-unit conflicts (ingest continued):")
        for warning in summary.unit_conflicts:
            print(f"- {warning}")


def command_extract(args: argparse.Namespace) -> int:
    summary = extract_pv_xlsx(
        xlsx_path=args.xlsx,
        output_dir=args.output_dir,
        dataset_prefix=args.dataset_prefix,
    )
    print_extract_summary(summary)
    return 0


def command_load(args: argparse.Namespace) -> int:
    dsn = load_env_dsn(args.dsn)
    summary = run_load_pv(
        path=args.path,
        dsn=dsn,
        source_url=args.source_url,
        dataset_prefix_override=args.dataset_prefix,
        force_reporting_year_override=args.force_reporting_year,
    )
    print_load_summary(summary)
    return 0


def command_run(args: argparse.Namespace) -> int:
    extract_summary = extract_pv_xlsx(
        xlsx_path=args.xlsx,
        output_dir=args.output_dir,
        dataset_prefix=args.dataset_prefix,
    )
    print_extract_summary(extract_summary)

    dsn = load_env_dsn(args.dsn)
    load_summary = run_load_pv_paths(
        csv_paths=[sheet.csv_path for sheet in extract_summary.sheets],
        dsn=dsn,
        source_url=args.source_url,
        dataset_prefix_override=args.dataset_prefix,
    )
    print_load_summary(load_summary)
    return 0


def add_extract_parser(subparsers: SubparserAction) -> None:
    parser = subparsers.add_parser(
        "extract-pv-xlsx",
        help="Extract visible pivot sheets from PivotTableViewer workbook into normalized CSVs",
    )
    parser.add_argument("--xlsx", required=True, type=Path, help="Path to PivotTableViewer workbook")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"Output directory for normalized CSVs (default: {DEFAULT_OUTPUT_DIR})",
    )
    parser.add_argument(
        "--dataset-prefix",
        required=True,
        help="Dataset prefix (expected NAEI2024pv)",
    )
    parser.set_defaults(func=command_extract)


def add_load_parser(subparsers: SubparserAction) -> None:
    parser = subparsers.add_parser("load-pv", help="Load normalized PV CSV(s) into Supabase/Postgres")
    parser.add_argument("--path", required=True, type=Path, help="CSV file or directory containing CSV files")
    parser.add_argument("--dsn", help="Database DSN (optional if SUPABASE_DB_URL/DATABASE_URL set)")
    parser.add_argument("--source-url", help="Optional source URL to store in dataset_file")
    parser.add_argument("--dataset-prefix", help="Override dataset_prefix from CSV rows")
    parser.add_argument(
        "--force-reporting-year",
        type=parse_year_arg,
        help="Optional override for reporting_year while loading CSV rows",
    )
    parser.set_defaults(func=command_load)


def add_run_parser(subparsers: SubparserAction) -> None:
    parser = subparsers.add_parser("run-pv-ingest", help="Extract + load PV workbook in one command")
    parser.add_argument("--xlsx", required=True, type=Path, help="Path to PivotTableViewer workbook")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"Output directory for normalized CSVs (default: {DEFAULT_OUTPUT_DIR})",
    )
    parser.add_argument("--dsn", help="Database DSN (optional if SUPABASE_DB_URL/DATABASE_URL set)")
    parser.add_argument("--dataset-prefix", required=True, help="Dataset prefix (expected NAEI2024pv)")
    parser.add_argument("--source-url", help="Optional source URL to store in dataset_file")
    parser.set_defaults(func=command_run)


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="NAEI PV extractor/loader")
    subparsers = parser.add_subparsers(dest="command", required=True)
    add_extract_parser(subparsers)
    add_load_parser(subparsers)
    add_run_parser(subparsers)

    args = parser.parse_args(argv)
    try:
        return int(args.func(args))
    except Exception as exc:  # pragma: no cover - CLI top-level handler
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
